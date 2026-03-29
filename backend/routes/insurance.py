from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from database import get_db
import models
import auth as auth_utils
import openpyxl
import io
from data.seed_data import SAL_HABRIUT_COVERAGES, HMO_COVERAGES

router = APIRouter(prefix="/api/patients/{patient_id}/insurance", tags=["insurance"])


class InsuranceSourceCreate(BaseModel):
    source_type: str
    hmo_name: Optional[str] = None
    hmo_level: Optional[str] = None
    company_name: Optional[str] = None
    policy_number: Optional[str] = None
    policy_type: Optional[str] = None
    notes: Optional[str] = None


class CoverageCreate(BaseModel):
    category: str
    is_covered: bool = True
    coverage_amount: Optional[float] = None
    coverage_percentage: Optional[float] = None
    copay: Optional[float] = None
    annual_limit: Optional[float] = None
    conditions: Optional[str] = None
    abroad_covered: bool = False
    notes: Optional[str] = None


class EntitlementCreate(BaseModel):
    entitlement_type: str
    title: str
    description: Optional[str] = None
    amount: Optional[float] = None
    is_approved: bool = False
    notes: Optional[str] = None


def source_to_dict(s):
    return {
        "id": s.id,
        "patient_id": s.patient_id,
        "source_type": s.source_type,
        "hmo_name": s.hmo_name,
        "hmo_level": s.hmo_level,
        "company_name": s.company_name,
        "policy_number": s.policy_number,
        "policy_type": s.policy_type,
        "notes": s.notes,
        "is_active": s.is_active,
        "coverages": [coverage_to_dict(c) for c in s.coverages],
    }


def coverage_to_dict(c):
    return {
        "id": c.id,
        "category": c.category,
        "is_covered": c.is_covered,
        "coverage_amount": c.coverage_amount,
        "coverage_percentage": c.coverage_percentage,
        "copay": c.copay,
        "annual_limit": c.annual_limit,
        "conditions": c.conditions,
        "abroad_covered": c.abroad_covered,
        "notes": c.notes,
    }


def get_patient_for_manager(patient_id: int, current_user: models.User, db: Session):
    patient = db.query(models.Patient).filter(models.Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    if current_user.role == "manager" and patient.manager_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    return patient


# ── Insurance Sources ──────────────────────────────────────

@router.get("")
def list_sources(patient_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_utils.get_current_user)):
    sources = db.query(models.InsuranceSource).filter(models.InsuranceSource.patient_id == patient_id).all()
    return [source_to_dict(s) for s in sources]


@router.post("")
def create_source(patient_id: int, data: InsuranceSourceCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_utils.require_manager)):
    get_patient_for_manager(patient_id, current_user, db)
    source = models.InsuranceSource(**data.model_dump(), patient_id=patient_id)
    db.add(source)
    db.flush()

    # Auto-populate coverages for סל הבריאות
    if data.source_type == "sal_habriut":
        for category, cov in SAL_HABRIUT_COVERAGES.items():
            coverage = models.Coverage(
                insurance_source_id=source.id,
                category=category,
                is_covered=cov.get("is_covered", True),
                coverage_percentage=cov.get("coverage_percentage"),
                copay=cov.get("copay"),
                annual_limit=cov.get("annual_limit"),
                conditions=cov.get("conditions"),
                abroad_covered=cov.get("abroad_covered", False),
                notes=cov.get("notes"),
            )
            db.add(coverage)

    # Auto-populate coverages for קופת חולים (משלים / פרמיום)
    elif data.source_type == "kupat_holim" and data.hmo_level in HMO_COVERAGES:
        for category, cov in HMO_COVERAGES[data.hmo_level].items():
            coverage = models.Coverage(
                insurance_source_id=source.id,
                category=category,
                is_covered=cov.get("is_covered", True),
                coverage_percentage=cov.get("coverage_percentage"),
                coverage_amount=cov.get("coverage_amount"),
                copay=cov.get("copay"),
                annual_limit=cov.get("annual_limit"),
                conditions=cov.get("conditions"),
                abroad_covered=cov.get("abroad_covered", False),
                notes=cov.get("notes"),
            )
            db.add(coverage)

    db.commit()
    db.refresh(source)
    return source_to_dict(source)


@router.delete("/{source_id}")
def delete_source(patient_id: int, source_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_utils.require_manager)):
    source = db.query(models.InsuranceSource).filter(
        models.InsuranceSource.id == source_id,
        models.InsuranceSource.patient_id == patient_id
    ).first()
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")
    db.delete(source)
    db.commit()
    return {"message": "Source deleted"}


# ── Coverage per source ────────────────────────────────────

@router.post("/{source_id}/coverage")
def add_coverage(patient_id: int, source_id: int, data: CoverageCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_utils.require_manager)):
    source = db.query(models.InsuranceSource).filter(
        models.InsuranceSource.id == source_id,
        models.InsuranceSource.patient_id == patient_id
    ).first()
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")
    # Remove existing coverage for this category
    db.query(models.Coverage).filter(
        models.Coverage.insurance_source_id == source_id,
        models.Coverage.category == data.category
    ).delete()
    coverage = models.Coverage(**data.model_dump(), insurance_source_id=source_id)
    db.add(coverage)
    db.commit()
    db.refresh(coverage)
    return coverage_to_dict(coverage)


@router.delete("/{source_id}/coverage/{coverage_id}")
def delete_coverage(patient_id: int, source_id: int, coverage_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_utils.require_manager)):
    cov = db.query(models.Coverage).filter(
        models.Coverage.id == coverage_id,
        models.Coverage.insurance_source_id == source_id,
    ).first()
    if not cov:
        raise HTTPException(status_code=404, detail="Coverage not found")
    db.delete(cov)
    db.commit()
    return {"message": "deleted"}


# ── Excel upload for הר הביטוח ────────────────────────────

def _find_col(row_dict, *keys):
    """Find value by trying multiple possible Hebrew/English column names."""
    for key in keys:
        for col, val in row_dict.items():
            if col and key.strip().lower() in col.strip().lower():
                v = str(val).strip() if val is not None else ""
                if v and v.lower() != "none":
                    return v
    return ""


@router.post("/upload-excel/preview")
async def preview_har_habitua_excel(
    patient_id: int,
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth_utils.require_manager)
):
    """Preview Excel contents before importing — returns detected headers and rows."""
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else f"עמודה {i+1}" for i, cell in enumerate(ws[1])]
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
            if any(row):
                rows.append([str(v) if v is not None else "" for v in row])
        return {"headers": headers, "sample_rows": rows, "total_rows": ws.max_row - 1}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"שגיאה בקריאת הקובץ: {str(e)}")


@router.post("/upload-excel")
async def upload_har_habitua_excel(
    patient_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth_utils.require_manager)
):
    get_patient_for_manager(patient_id, current_user, db)
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        policies = []
        skipped = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            row_dict = dict(zip(headers, row))

            # Flexible column detection — try many name variations
            company = _find_col(row_dict,
                "חברה", "חברת ביטוח", "מבטח", "company", "insurer", "שם חברה")
            policy_number = _find_col(row_dict,
                "מספר פוליסה", "פוליסה", "מס פוליסה", "policy", "policy number", "policy_number")
            policy_type = _find_col(row_dict,
                "סוג ביטוח", "סוג", "type", "kind", "ענף", "מוצר")
            notes = _find_col(row_dict,
                "הערות", "הערה", "notes", "remarks", "comment")

            # Fallback: if no company found, use first non-empty cell
            if not company:
                for v in row:
                    if v and str(v).strip():
                        company = str(v).strip()
                        break

            if company:
                source = models.InsuranceSource(
                    patient_id=patient_id,
                    source_type="har_habitua",
                    company_name=company,
                    policy_number=policy_number or None,
                    policy_type=policy_type or "ביטוח רפואי",
                    notes=notes or None,
                )
                db.add(source)
                policies.append({
                    "row": row_idx,
                    "company": company,
                    "policy_number": policy_number,
                    "policy_type": policy_type,
                })
            else:
                skipped.append(row_idx)

        db.commit()
        return {
            "message": f"יובאו {len(policies)} פוליסות בהצלחה",
            "imported": len(policies),
            "skipped": len(skipped),
            "policies": policies,
            "detected_headers": headers,
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"שגיאה בניתוח קובץ Excel: {str(e)}")


# ── Entitlements (ביטוח לאומי) ────────────────────────────

entitlement_router = APIRouter(prefix="/api/patients/{patient_id}/entitlements", tags=["entitlements"])


def entitlement_to_dict(e):
    return {
        "id": e.id,
        "patient_id": e.patient_id,
        "entitlement_type": e.entitlement_type,
        "title": e.title,
        "description": e.description,
        "amount": e.amount,
        "is_approved": e.is_approved,
        "notes": e.notes,
        "created_at": str(e.created_at) if e.created_at else None,
    }


@entitlement_router.get("")
def list_entitlements(patient_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_utils.get_current_user)):
    entitlements = db.query(models.Entitlement).filter(models.Entitlement.patient_id == patient_id).all()
    return [entitlement_to_dict(e) for e in entitlements]


@entitlement_router.post("")
def create_entitlement(patient_id: int, data: EntitlementCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_utils.require_manager)):
    entitlement = models.Entitlement(**data.model_dump(), patient_id=patient_id)
    db.add(entitlement)
    db.commit()
    db.refresh(entitlement)
    return entitlement_to_dict(entitlement)


@entitlement_router.put("/{entitlement_id}")
def update_entitlement(patient_id: int, entitlement_id: int, data: EntitlementCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_utils.require_manager)):
    e = db.query(models.Entitlement).filter(models.Entitlement.id == entitlement_id, models.Entitlement.patient_id == patient_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Entitlement not found")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(e, field, value)
    db.commit()
    db.refresh(e)
    return entitlement_to_dict(e)


@entitlement_router.delete("/{entitlement_id}")
def delete_entitlement(patient_id: int, entitlement_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_utils.require_manager)):
    e = db.query(models.Entitlement).filter(models.Entitlement.id == entitlement_id, models.Entitlement.patient_id == patient_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Entitlement not found")
    db.delete(e)
    db.commit()
    return {"message": "Deleted"}
