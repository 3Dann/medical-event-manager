import json
import io
import logging
import threading
import uuid
import time
import unicodedata

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from database import get_db, SessionLocal
import models
import auth as auth_utils
from doctor_normalize import normalize_record, fix_rtl_parens as _fix_rtl_parens

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ── In-memory Excel import job registry ───────────────────────────────────────
_import_jobs: dict = {}

def _cleanup_old_jobs():
    cutoff = time.time() - 3600
    for jid in list(_import_jobs):
        if _import_jobs[jid].get("started_at", 0) < cutoff:
            del _import_jobs[jid]

logger = logging.getLogger("doctors")

router = APIRouter(prefix="/api/doctors", tags=["doctors"])

HMO_OPTIONS = ["clalit", "maccabi", "meuhedet", "leumit"]


class DoctorCreate(BaseModel):
    name: str
    title: Optional[str] = None
    specialty: Optional[str] = None
    sub_specialty: Optional[str] = None
    license_number: Optional[str] = None
    phone: Optional[str] = None
    phone2: Optional[str] = None
    whatsapp: Optional[str] = None
    email: Optional[str] = None
    city: Optional[str] = None
    location: Optional[str] = None
    private_price: Optional[int] = None
    hmo_acceptance: Optional[List[str]] = None
    gives_expert_opinion: bool = False
    notes: Optional[str] = None
    extra_data: Optional[str] = None   # JSON string
    source_url: Optional[str] = None


class DoctorUpdate(BaseModel):
    name: Optional[str] = None
    title: Optional[str] = None
    specialty: Optional[str] = None
    sub_specialty: Optional[str] = None
    license_number: Optional[str] = None
    phone: Optional[str] = None
    phone2: Optional[str] = None
    whatsapp: Optional[str] = None
    email: Optional[str] = None
    city: Optional[str] = None
    location: Optional[str] = None
    private_price: Optional[int] = None
    hmo_acceptance: Optional[List[str]] = None
    gives_expert_opinion: Optional[bool] = None
    notes: Optional[str] = None
    extra_data: Optional[str] = None   # JSON string
    source_url: Optional[str] = None


def doctor_to_dict(d: models.Doctor) -> dict:
    hmo = []
    if d.hmo_acceptance:
        try:
            hmo = json.loads(d.hmo_acceptance)
        except Exception:
            hmo = []
    extra = {}
    if getattr(d, 'extra_data', None):
        try:
            extra = json.loads(d.extra_data)
        except Exception:
            extra = {}
    return {
        "id":                   d.id,
        "title":                getattr(d, 'title', None),
        "name":                 d.name,
        "specialty":            d.specialty,
        "sub_specialty":        d.sub_specialty,
        "license_number":       getattr(d, 'license_number', None),
        "phone":                d.phone,
        "phone2":               getattr(d, 'phone2', None),
        "whatsapp":             getattr(d, 'whatsapp', None),
        "email":                getattr(d, 'email', None),
        "city":                 getattr(d, 'city', None),
        "location":             d.location,
        "private_price":        getattr(d, 'private_price', None),
        "hmo_acceptance":       hmo,
        "gives_expert_opinion": d.gives_expert_opinion,
        "notes":                d.notes,
        "extra_data":           extra,
        "source_url":           d.source_url,
        "created_at":           str(d.created_at) if d.created_at else None,
    }


@router.get("/schema")
def get_doctor_schema(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth_utils.get_current_user),
):
    fixed = ["name","license_number","specialty","sub_specialty","phone","phone2",
             "whatsapp","email","city","location","private_price",
             "hmo_acceptance","gives_expert_opinion","notes"]
    extra_keys: set = set()
    for row in db.query(models.Doctor.extra_data).filter(
        models.Doctor.extra_data.isnot(None)
    ).limit(300).all():
        try:
            extra_keys.update(json.loads(row.extra_data).keys())
        except Exception:
            pass
    return {"fixed": fixed, "extra": sorted(extra_keys)}


@router.get("/filter-options")
def get_filter_options(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth_utils.get_current_user),
):
    """Return distinct values for specialty, sub_specialty, and location dropdowns."""
    def distinct(col):
        rows = db.query(col).filter(col.isnot(None), col != "").distinct().order_by(col).all()
        return [r[0] for r in rows]

    specialties = distinct(models.Doctor.specialty)
    sub_specialties = distinct(models.Doctor.sub_specialty)

    # For location extract city/area (last comma-separated token or whole value)
    raw_locations = distinct(models.Doctor.location)
    areas = sorted({loc.split(",")[-1].strip() for loc in raw_locations if loc})

    return {"specialties": specialties, "sub_specialties": sub_specialties, "areas": areas}


@router.get("/export/excel")
def export_doctors_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth_utils.get_current_user),
):
    """Export full doctors database as RTL Excel file."""
    from fastapi.responses import StreamingResponse
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl לא מותקן")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "מאגר רופאים"
    ws.sheet_view.rightToLeft = True  # RTL sheet

    # Header style
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2563EB")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["שם הרופא", "מומחיות", "תת-מומחיות", "טלפון", "טלפון 2", "וואטסאפ", "אימייל", "עיר", "מיקום", "מחיר פרטי", "קופות חולים", "חוות דעת", "הערות", "מקור"]
    col_widths = [22, 18, 18, 16, 16, 16, 24, 14, 22, 12, 24, 12, 28, 30]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Alternate row fill
    fill_even = PatternFill("solid", fgColor="EFF6FF")
    fill_odd = PatternFill("solid", fgColor="FFFFFF")

    doctors = db.query(models.Doctor).order_by(models.Doctor.specialty, models.Doctor.name).all()

    for row_idx, doc in enumerate(doctors, start=2):
        # Parse HMO list
        try:
            hmo_list = json.loads(doc.hmo_acceptance) if doc.hmo_acceptance else []
            hmo_map = {"clalit": "כללית", "maccabi": "מכבי", "meuhedet": "מאוחדת", "leumit": "לאומית"}
            hmo_str = " | ".join(hmo_map.get(h, h) for h in hmo_list)
        except Exception:
            hmo_str = doc.hmo_acceptance or ""

        row_data = [
            doc.name or "",
            doc.specialty or "",
            doc.sub_specialty or "",
            doc.phone or "",
            getattr(doc, "phone2", None) or "",
            getattr(doc, "whatsapp", None) or "",
            getattr(doc, "email", None) or "",
            getattr(doc, "city", None) or "",
            doc.location or "",
            getattr(doc, "private_price", None) or "",
            hmo_str,
            "כן" if doc.gives_expert_opinion else "לא",
            doc.notes or "",
            doc.source_url or "",
        ]
        fill = fill_even if row_idx % 2 == 0 else fill_odd
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = right_align
            cell.fill = fill
            cell.border = border

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''%D7%9E%D7%90%D7%92%D7%A8%20%D7%A8%D7%95%D7%A4%D7%90%D7%99%D7%9D.xlsx"},
    )


@router.get("")
def list_doctors(
    search: Optional[str] = None,
    specialty: Optional[str] = None,
    sub_specialty: Optional[str] = None,
    hmo: Optional[str] = None,
    location: Optional[str] = None,
    expert_opinion: Optional[bool] = None,
    limit: int = Query(default=500, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth_utils.get_current_user),
):
    q = db.query(models.Doctor)
    if search:
        q = q.filter(
            models.Doctor.name.ilike(f"%{search}%") |
            models.Doctor.specialty.ilike(f"%{search}%") |
            models.Doctor.sub_specialty.ilike(f"%{search}%") |
            models.Doctor.location.ilike(f"%{search}%") |
            models.Doctor.city.ilike(f"%{search}%") |
            models.Doctor.license_number.ilike(f"%{search}%")
        )
    if specialty:
        q = q.filter(models.Doctor.specialty.ilike(f"%{specialty}%"))
    if sub_specialty:
        q = q.filter(models.Doctor.sub_specialty.ilike(f"%{sub_specialty}%"))
    if hmo:
        q = q.filter(models.Doctor.hmo_acceptance.ilike(f"%{hmo}%"))
    if location:
        q = q.filter(models.Doctor.location.ilike(f"%{location}%"))
    if expert_opinion is not None:
        q = q.filter(models.Doctor.gives_expert_opinion == expert_opinion)
    total = q.count()
    doctors = q.order_by(models.Doctor.name).offset(offset).limit(limit).all()
    return {"total": total, "offset": offset, "limit": limit, "items": [doctor_to_dict(d) for d in doctors]}


@router.post("")
def create_doctor(
    data: DoctorCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth_utils.require_manager),
):
    rec = {
        "name":                 data.name,
        "title":                data.title,
        "specialty":            data.specialty,
        "sub_specialty":        data.sub_specialty,
        "license_number":       data.license_number,
        "phone":                data.phone,
        "phone2":               data.phone2,
        "whatsapp":             data.whatsapp,
        "email":                data.email,
        "city":                 data.city,
        "location":             data.location,
        "private_price":        data.private_price,
        "hmo_acceptance":       json.dumps(data.hmo_acceptance or [], ensure_ascii=False),
        "gives_expert_opinion": data.gives_expert_opinion,
        "notes":                data.notes,
        "extra_data":           data.extra_data,
        "source_url":           data.source_url,
    }
    rec = normalize_record(rec)
    if rec is None:
        raise HTTPException(status_code=400, detail="שם הרופא אינו תקין")
    doctor = models.Doctor(**rec)
    db.add(doctor)
    db.commit()
    db.refresh(doctor)
    return doctor_to_dict(doctor)


@router.put("/{doctor_id}")
def update_doctor(
    doctor_id: int,
    data: DoctorUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth_utils.require_manager),
):
    doctor = db.query(models.Doctor).filter(models.Doctor.id == doctor_id).first()
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    update_data = data.model_dump(exclude_none=True)
    if "hmo_acceptance" in update_data:
        update_data["hmo_acceptance"] = json.dumps(update_data["hmo_acceptance"], ensure_ascii=False)
    for field, value in update_data.items():
        setattr(doctor, field, value)
    db.commit()
    db.refresh(doctor)
    return doctor_to_dict(doctor)


@router.delete("/all")
def delete_all_doctors(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth_utils.require_manager),
):
    deleted = db.query(models.Doctor).delete()
    db.commit()
    return {"deleted": deleted}


@router.delete("/{doctor_id}")
def delete_doctor(
    doctor_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth_utils.require_manager),
):
    doctor = db.query(models.Doctor).filter(models.Doctor.id == doctor_id).first()
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    db.delete(doctor)
    db.commit()
    return {"message": "Doctor deleted"}


# ── Excel Import ───────────────────────────────────────────────────────────────

def _normalize_name(name: str) -> str:
    if not name:
        return ""
    import re
    n = unicodedata.normalize("NFC", name.strip().lower())
    return re.sub(r"\s+", " ", n)


def _clean_cell(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    BIDI_CONTROLS = {
        '​', '‌', '‍', '‎', '‏',
        '‪', '‫', '‬', '‭', '‮',
        '⁦', '⁧', '⁨', '⁩', '﻿',
    }
    text = ''.join(c for c in text if c not in BIDI_CONTROLS)
    import re
    text = re.sub(r' {2,}', ' ', text).strip()
    text = _fix_rtl_parens(text)
    return text


def _parse_hmo_string(value: str) -> List[str]:
    result = []
    val = str(value).lower()
    hmo_map = {
        "clalit":   ["כללית", "clalit"],
        "maccabi":  ["מכבי", "maccabi"],
        "meuhedet": ["מאוחדת", "meuhedet"],
        "leumit":   ["לאומית", "leumit"],
    }
    for key, aliases in hmo_map.items():
        if any(alias in val for alias in aliases):
            result.append(key)
    if not result and ("כן" in val or "yes" in val or "all" in val or "כל" in val):
        result = ["clalit", "maccabi", "meuhedet", "leumit"]
    return result


def _bool_from_value(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    val = str(value).strip().lower()
    return val in ["כן", "yes", "true", "1", "נכון"]


_FIELD_ALIASES = {
    "name":                ["שם הרופא", "שם רופא", "שם", "name", "doctor", "full name"],
    "title":               ["תואר", "title", "degree"],
    "specialty":           ["מומחיות", "specialty", "התמחות", "תחום"],
    "sub_specialty":       ["תת-מומחיות", "תת מומחיות", "תת-התמחות", "תת התמחות", "sub_specialty"],
    "license_number":      ["מספר רישיון", "רישיון", "license", "מס רישיון"],
    "phone":               ["טלפון ראשי", "טלפון", "phone", "נייד", "mobile", "tel"],
    "phone2":              ["טלפון 2", "טלפון נוסף", "טלפון שני", "מזכירה", "phone2"],
    "whatsapp":            ["וואטסאפ", "whatsapp", "ווצאפ"],
    "email":               ["אימייל", "מייל", "email", "e-mail"],
    "city":                ["עיר", "city", "ישוב"],
    "location":            ["מיקום", "כתובת", "location", "היכן מקבל", "קליניקה", "address"],
    "private_price":       ["מחיר פרטי", "תשלום", "עלות", "מחיר", "price"],
    "hmo_acceptance":      ["קופות חולים", "קופה", "hmo", "חברת ביטוח", "ביטוח"],
    "languages":           ["שפות", "languages", "שפה"],
    "gives_expert_opinion":["חוות דעת", "ועדות", "expert_opinion", "opinion"],
    "notes":               ["הערות", "notes", "מידע נוסף"],
}


def _run_import_job(content: bytes, job_id: str, field_aliases: dict):
    job = _import_jobs[job_id]
    db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        total = max(ws.max_row - 1, 0)
        job["total"] = total

        raw_headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        headers_lc  = [h.lower() for h in raw_headers]

        col_map = {}
        for field, aliases in field_aliases.items():
            for i, h in enumerate(headers_lc):
                if h and any(alias.lower() in h for alias in aliases):
                    col_map[field] = i
                    break

        if "name" not in col_map:
            found = ", ".join(h for h in raw_headers if h) or "אין"
            job["status"]  = "error"
            job["message"] = f"לא זוהתה עמודת שם. עמודות: {found}"
            return

        job["detected_columns"] = list(col_map.keys())
        mapped_indices = set(col_map.values())
        unmapped_cols = {raw_headers[i]: i for i, h in enumerate(raw_headers)
                         if h and i not in mapped_indices}

        def get_cell(row, field):
            idx = col_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        existing = {_normalize_name(d.name) for d in db.query(models.Doctor.name).all() if d.name}
        imported = skipped_dup = skipped_inv = 0
        skip_samples, row_errors = [], []

        def _skip(raw, reason):
            if len(skip_samples) < 5:
                skip_samples.append({"name": str(raw)[:60], "reason": reason})

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(v for v in row if v is not None):
                continue
            try:
                def gc(field):
                    return _clean_cell(get_cell(row, field))

                raw_name = get_cell(row, "name")
                name = _clean_cell(raw_name)
                if not name or name.lower() in ("none", "nan"):
                    _skip(raw_name, "שם ריק"); skipped_inv += 1; continue

                norm = _normalize_name(name)
                if norm in existing:
                    skipped_dup += 1; continue

                title_val  = gc("title") or None
                spec_raw   = gc("specialty")
                spec_parts = [s.strip() for s in spec_raw.split(",") if s.strip()]
                sub_col    = gc("sub_specialty") or None
                lang_val   = gc("languages")
                notes_val  = gc("notes")
                notes_comb = " | ".join(filter(None, [
                    f"שפות: {lang_val}" if lang_val else "", notes_val])) or None

                hmo_raw   = get_cell(row, "hmo_acceptance")
                raw_price = get_cell(row, "private_price")
                try:
                    parsed_price = int(float(str(raw_price).replace(",","").strip())) if raw_price else None
                except (ValueError, TypeError):
                    parsed_price = None

                lic_raw = gc("license_number") or None
                rec = {
                    "name":                 name,
                    "title":                title_val,
                    "specialty":            spec_parts[0] if spec_parts else None,
                    "sub_specialty":        sub_col or (spec_parts[1] if len(spec_parts) > 1 else None),
                    "license_number":       lic_raw,
                    "phone":                gc("phone")    or None,
                    "phone2":               gc("phone2")   or None,
                    "whatsapp":             gc("whatsapp") or None,
                    "email":                gc("email")    or None,
                    "city":                 gc("city")     or None,
                    "location":             gc("location") or None,
                    "private_price":        parsed_price,
                    "hmo_acceptance":       json.dumps(_parse_hmo_string(hmo_raw) if hmo_raw else [], ensure_ascii=False),
                    "gives_expert_opinion": _bool_from_value(get_cell(row, "gives_expert_opinion")),
                    "notes":                notes_comb,
                    "source_url":           "excel_import",
                    "extra_data":           json.dumps(
                        {col: str(row[idx] or "").strip()
                         for col, idx in unmapped_cols.items()
                         if idx < len(row) and row[idx] is not None and str(row[idx]).strip()},
                        ensure_ascii=False) or None,
                }
                rec = normalize_record(rec)
                if rec is None:
                    _skip(name, "שם לא תקין"); skipped_inv += 1; continue

                db.add(models.Doctor(**rec))
                existing.add(norm)
                imported += 1

                if imported % 500 == 0:
                    db.commit()
                    job["imported"] = imported
                    job["skipped_duplicates"] = skipped_dup
                    job["skipped_invalid"] = skipped_inv

            except Exception as e:
                row_errors.append(f"שורה {row_num}: {e}")
                _skip(f"שורה {row_num}", f"שגיאה: {e}")
                skipped_inv += 1

        try:
            db.commit()
        except Exception as e:
            db.rollback()
            job["status"]  = "error"
            job["message"] = f"שגיאת שמירה: {e}"
            return

        job.update({
            "status":             "done",
            "imported":           imported,
            "skipped_duplicates": skipped_dup,
            "skipped_invalid":    skipped_inv,
            "skip_samples":       skip_samples,
            "errors":             row_errors[:5],
            "message":            f"הסתיים — יובאו {imported:,} רופאים",
        })

    except Exception as e:
        job["status"]  = "error"
        job["message"] = str(e)
    finally:
        db.close()


@router.get("/import/status/{job_id}")
async def get_import_status(
    job_id: str,
    current_user: models.User = Depends(auth_utils.require_manager),
):
    job = _import_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


@router.post("/import/excel")
async def import_from_excel(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth_utils.require_manager),
):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl לא מותקן")

    try:
        content = await file.read()
        openpyxl.load_workbook(io.BytesIO(content), read_only=True).close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"לא ניתן לפתוח את הקובץ: {e}")

    _cleanup_old_jobs()
    job_id = uuid.uuid4().hex[:10]
    _import_jobs[job_id] = {
        "status":             "running",
        "imported":           0,
        "skipped_duplicates": 0,
        "skipped_invalid":    0,
        "total":              0,
        "message":            "מתחיל...",
        "started_at":         time.time(),
    }

    t = threading.Thread(target=_run_import_job, args=(content, job_id, _FIELD_ALIASES), daemon=True)
    t.start()

    return {"job_id": job_id}
